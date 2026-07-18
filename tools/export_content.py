"""Valida Excel y genera los JSON de contenido y traducciones."""
from __future__ import annotations
import argparse, json, re, sys
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

SOURCE_PATH = Path(__file__).resolve()
ROOT = SOURCE_PATH.parents[1] if SOURCE_PATH.parent.name == "tools" else SOURCE_PATH.parents[2]
DEFAULT_BOOK = ROOT / "data/design/grimorium_content.xlsx"
CATEGORIES = {"potion":{"consumable"}, "power":{"element","shot_modifier"}, "book":{"element","combo"}}
OPERATIONS = {"add","set","multiply"}
SHOT_FIELDS = {"fan_count","fan_angle","parallel_count","parallel_spacing","radius_multiplier","damage_multiplier","speed_multiplier","range_multiplier","cooldown_multiplier","rear_shot"}
PLACEHOLDER = re.compile(r"\{(?:good:|bad:)?([a-zA-Z_][a-zA-Z0-9_]*)\}")

class ContentError(Exception): pass

def rows(book, name):
    if name not in book.sheetnames: raise ContentError(f"Falta la hoja {name}")
    raw = list(book[name].iter_rows(values_only=True))
    if not raw: return []
    headers = [str(v).strip() if v is not None else "" for v in raw[0]]
    if not all(headers): raise ContentError(f"{name}: cabecera incompleta")
    result=[]
    for number, values in enumerate(raw[1:], 2):
        if any(v not in (None,"") for v in values):
            row=dict(zip(headers,values)); row["_row"]=number; result.append(row)
    return result

def as_bool(value, where):
    if isinstance(value,bool): return value
    if isinstance(value,(int,float)) and value in (0,1): return bool(value)
    if isinstance(value,str) and value.lower().strip() in {"true","false","yes","no","si","sí"}: return value.lower().strip() in {"true","yes","si","sí"}
    raise ContentError(f"{where}: enabled debe ser TRUE/FALSE")

def clean(row): return {k:v for k,v in row.items() if k!="_row" and v not in (None,"")}

def typed_value(row, where):
    kind,value=str(row.get("value_type","number")),row.get("value")
    if kind=="bool": return as_bool(value,where)
    if kind=="int" and isinstance(value,(int,float)) and not isinstance(value,bool) and int(value)==value: return int(value)
    if kind=="number" and isinstance(value,(int,float)) and not isinstance(value,bool): return value
    raise ContentError(f"{where}: value no coincide con value_type={kind}")

def nest(flat):
    root={}
    for key,value in sorted(flat.items()):
        parts=key.split("."); cursor=root
        for part in parts[:-1]: cursor=cursor.setdefault(part,{})
        cursor[parts[-1]]=value
    return root

def validate(path):
    book=load_workbook(path,data_only=False,read_only=True)
    item_rows,effect_rows,req_rows,text_rows,detail_rows=(rows(book,n) for n in ("Items","Effects","Requirements","Texts","Detail Lines"))
    errors=[]; items=[]; effects=[]; requirements=[]; item_keys=set(); effect_keys=set(); fields=defaultdict(set)
    for row in item_rows:
        where=f"Items fila {row['_row']}"
        try:
            typ,item_id,category=(str(row.get(k,"")).strip() for k in ("type","id","category")); key=(typ,item_id)
            if typ not in CATEGORIES or category not in CATEGORIES[typ]: raise ContentError(f"{where}: tipo/categoria invalidos")
            if not re.fullmatch(r"[a-z0-9_]+",item_id): raise ContentError(f"{where}: id invalido")
            if key in item_keys: raise ContentError(f"{where}: id duplicado {typ}/{item_id}")
            enabled=as_bool(row.get("enabled"),where); price=row.get("price"); chance=row.get("chance"); asset=str(row.get("asset","")).replace("\\","/")
            if isinstance(price,bool) or not isinstance(price,(int,float)) or price<0: raise ContentError(f"{where}: price invalido")
            if chance not in (None,"") and (not isinstance(chance,(int,float)) or isinstance(chance,bool) or not 0<=chance<=1): raise ContentError(f"{where}: chance fuera de 0..1")
            if enabled and not (ROOT/"assets"/asset).is_file(): raise ContentError(f"{where}: falta assets/{asset}")
            if not row.get("text_key"): raise ContentError(f"{where}: falta text_key")
            item_keys.add(key); value=clean(row); value.update(type=typ,id=item_id,category=category,enabled=enabled,price=price,asset=asset); items.append(value)
        except (ContentError,TypeError,ValueError) as exc: errors.append(str(exc))
    for row in effect_rows:
        where=f"Effects fila {row['_row']}"
        try:
            owner=(str(row.get("owner_type","")),str(row.get("owner_id",""))); target=str(row.get("target","")); key=(*owner,target)
            if owner not in item_keys: raise ContentError(f"{where}: owner inexistente {owner}")
            if key in effect_keys: raise ContentError(f"{where}: efecto duplicado {key}")
            if row.get("operation") not in OPERATIONS: raise ContentError(f"{where}: operation invalida")
            if row.get("target_scope")=="shot_context" and target not in SHOT_FIELDS: raise ContentError(f"{where}: campo de disparo desconocido {target}")
            effect_keys.add(key); fields[owner].add(target); value=clean(row); value["value"]=typed_value(row,where); effects.append(value)
        except (ContentError,TypeError,ValueError) as exc: errors.append(str(exc))
    seen=set()
    for row in req_rows:
        where=f"Requirements fila {row['_row']}"; owner=(str(row.get("owner_type","")),str(row.get("owner_id",""))); req_id=str(row.get("requirement_id","")); key=(*owner,str(row.get("requirement_type","")),req_id)
        if owner not in item_keys or row.get("requirement_type")!="power_owned" or ("power",req_id) not in item_keys: errors.append(f"{where}: requisito invalido")
        elif key in seen: errors.append(f"{where}: requisito duplicado")
        else: seen.add(key); requirements.append(clean(row))
    flat={"es":{},"en":{}}
    for row in text_rows:
        key=str(row.get("key",""))
        for language in flat:
            if not key or row.get(language) in (None,""): errors.append(f"Texts fila {row['_row']}: falta {language}/{key}")
            else: flat[language][key]=str(row[language])
    seen=set()
    for row in detail_rows:
        language,key,order,line=str(row.get("language","")),str(row.get("text_key","")),row.get("order"),row.get("text"); marker=(language,key,order)
        if language not in flat or not key or not isinstance(order,int) or line in (None,"") or marker in seen: errors.append(f"Detail Lines fila {row['_row']}: fila invalida/duplicada")
        else: seen.add(marker); flat[language].setdefault(f"{key}.detail",[]).append((order,str(line)))
    for language in flat:
        for key,value in list(flat[language].items()):
            if isinstance(value,list): flat[language][key]=[line for _,line in sorted(value)]
    derived={"slow_percent","chain_damage_percent","damage_received_percent","fragment_damage_percent","execute_stack_percent","chance_percent"}
    for item in (i for i in items if i["enabled"]):
        owner=(item["type"],item["id"]); available=set(item)|fields[owner]|derived
        if item["type"]=="potion": available.add("amount")
        for language in flat:
            for suffix in ("name","short"):
                key=f"{item['text_key']}.{suffix}"
                if key not in flat[language]: errors.append(f"Falta traduccion {language}: {key}")
            for line in flat[language].get(f"{item['text_key']}.detail",[]):
                unknown=set(PLACEHOLDER.findall(line))-available
                if unknown: errors.append(f"{language}/{item['text_key']}: placeholders desconocidos {sorted(unknown)}")
    for typ in CATEGORIES:
        prices={i["price"] for i in items if i["type"]==typ and i["enabled"]}
        if len(prices)>1: errors.append(f"Los precios activos de {typ} deben coincidir: {sorted(prices)}")
    if errors: raise ContentError("\n".join(f"- {e}" for e in errors))
    return items,effects,requirements,{language:nest(values) for language,values in flat.items()}

def write_json(path,value):
    path.parent.mkdir(parents=True,exist_ok=True)
    path.write_text(json.dumps(value,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")

def main():
    parser=argparse.ArgumentParser()
    parser.add_argument("workbook",nargs="?",type=Path,default=DEFAULT_BOOK)
    parser.add_argument("--check",action="store_true")
    args=parser.parse_args()
    try: items,effects,requirements,translations=validate(args.workbook.resolve())
    except (ContentError,FileNotFoundError) as exc: print(f"ERROR DE CONTENIDO:\n{exc}",file=sys.stderr); return 1
    if args.check: print(f"OK: {len(items)} objetos, {len(effects)} efectos y {len(requirements)} requisitos"); return 0
    for name,value in (("items",items),("effects",effects),("requirements",requirements)): write_json(ROOT/"data/game"/f"{name}.json",value)
    for language,value in translations.items(): write_json(ROOT/"data/lang"/f"{language}.json",value)
    print(f"Exportado correctamente desde {args.workbook}"); return 0

if __name__=="__main__": raise SystemExit(main())
